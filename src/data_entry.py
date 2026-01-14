import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os



# Configuration
DEFAULT_EXCEL_FILE = r"sample.xlsx"

# Column definitions - will be set dynamically
COLUMNS = []
DISPLAY_COLUMNS = []


class Config:
    """Configuration manager with multi-sheet support"""
    def __init__(self):
        self.excel_path = None
        self.sheets = {}  # {sheet_name: {'columns': [], 'display_columns': []}}
        self.load_config()
    
    def load_config(self):
        """Load or prompt for Excel file location"""
        if os.path.exists(DEFAULT_EXCEL_FILE):
            self.excel_path = DEFAULT_EXCEL_FILE
            # Load sheets and columns from existing file
            self._load_from_existing_file()
        else:
            # File doesn't exist - let user create it
            print("\n" + "="*60)
            print("EXCEL FILE NOT FOUND - SETUP REQUIRED")
            print("="*60)
            
            create_new = input(f"\n'{DEFAULT_EXCEL_FILE}' does not exist.\nWould you like to create it? (yes/no): ").strip().lower()
            
            if create_new in ['yes', 'y']:
                self._setup_new_file()
            else:
                # Ask to select existing file
                use_existing = input("\nWould you like to select an existing Excel file? (yes/no): ").strip().lower()
                if use_existing in ['yes', 'y']:
                    root = tk.Tk()
                    root.withdraw()
                    path = filedialog.askopenfilename(
                        title="Select Excel File",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                    )
                    root.destroy()
                    
                    if path:
                        self.excel_path = path
                        self._load_from_existing_file()
                    else:
                        print("\nNo file selected. Exiting...")
                else:
                    print("\nNo file selected. Exiting...")
    
    def _setup_new_file(self):
        """Interactive setup for new Excel file with multiple sheets"""
        print("\n" + "="*60)
        print("SHEET SETUP")
        print("="*60)
        
        # Ask how many sheets
        while True:
            try:
                num_sheets = input("\nHow many sheets do you want to create? (1-10): ").strip()
                num_sheets = int(num_sheets)
                if 1 <= num_sheets <= 10:
                    break
                else:
                    print("⚠ Please enter a number between 1 and 10.")
            except ValueError:
                print("⚠ Invalid input. Please enter a number.")
        
        print(f"\n✓ You will create {num_sheets} sheet(s).\n")
        
        # Setup each sheet
        for sheet_num in range(1, num_sheets + 1):
            print("="*60)
            print(f"SETTING UP SHEET {sheet_num} of {num_sheets}")
            print("="*60)
            
            # Get sheet name
            while True:
                sheet_name = input(f"\nEnter name for Sheet {sheet_num}: ").strip()
                if not sheet_name:
                    print("⚠ Sheet name cannot be empty.")
                    continue
                if sheet_name in self.sheets:
                    print(f"⚠ Sheet '{sheet_name}' already exists. Choose a different name.")
                    continue
                break
            
            print(f"\n✓ Sheet name: {sheet_name}")
            
            # Setup columns for this sheet
            columns = self._setup_columns_for_sheet(sheet_name)
            display_columns = self._setup_display_columns(columns)
            
            self.sheets[sheet_name] = {
                'columns': columns,
                'display_columns': display_columns
            }
            
            print(f"\n✓ Sheet '{sheet_name}' configured successfully!")
            print(f"  - Columns: {len(columns)}")
            print(f"  - Preview columns: {len(display_columns)}")
        
        # Create the Excel file with all sheets
        self._create_excel_file()
    
    def _setup_columns_for_sheet(self, sheet_name):
        """Setup columns for a specific sheet"""
        print(f"\n--- Column Setup for '{sheet_name}' ---")
        print("Enter column names one by one. Press Enter with empty input to finish.")
        
        columns = []
        column_num = 1
        
        while True:
            col_name = input(f"  Column {column_num}: ").strip()
            
            if not col_name:
                if len(columns) == 0:
                    print("  ⚠ You must define at least one column!")
                    continue
                else:
                    break
            
            if col_name in columns:
                print(f"  ⚠ '{col_name}' already exists. Please use a different name.")
                continue
            
            columns.append(col_name)
            print(f"  ✓ Added: {col_name}")
            column_num += 1
        
        return columns
    
    def _setup_display_columns(self, columns):
        """Setup display columns for preview"""
        print(f"\nSelect columns to display in preview (Total available: {len(columns)}):")
        for idx, col in enumerate(columns, 1):
            print(f"  {idx}. {col}")
        
        print("\nEnter column numbers separated by commas (e.g., 1,2,3,4)")
        print("Or press Enter to use first 4 columns.")
        
        preview_input = input("\nPreview columns: ").strip()
        
        if preview_input:
            try:
                indices = [int(x.strip()) for x in preview_input.split(',')]
                display_columns = [columns[i-1] for i in indices if 1 <= i <= len(columns)]
                
                if not display_columns:
                    print("⚠ Invalid selection. Using first 4 columns.")
                    display_columns = columns[:min(4, len(columns))]
            except:
                print("⚠ Invalid input. Using first 4 columns.")
                display_columns = columns[:min(4, len(columns))]
        else:
            display_columns = columns[:min(4, len(columns))]
        
        print("✓ Preview columns:", ", ".join(display_columns))
        return display_columns
    
    def _create_excel_file(self):
        """Create Excel file with all configured sheets"""
        print("\n" + "="*60)
        print("CREATING EXCEL FILE")
        print("="*60)
        
        file_path = input(f"\nEnter file path (press Enter for '{DEFAULT_EXCEL_FILE}'): ").strip()
        
        if not file_path:
            file_path = DEFAULT_EXCEL_FILE
        
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
        
        try:
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # Create all sheets with headers
            for sheet_name, sheet_config in self.sheets.items():
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(sheet_config['columns'])
            
            workbook.save(file_path)
            
            self.excel_path = file_path
            print(f"\n✓ Successfully created: {file_path}")
            print(f"✓ Total sheets: {len(self.sheets)}")
            for sheet_name in self.sheets:
                print(f"  - {sheet_name}: {len(self.sheets[sheet_name]['columns'])} columns")
            
        except Exception as e:
            print(f"\n✗ Error creating file: {e}")
            print("Exiting...")
            self.excel_path = None
    
    def _load_from_existing_file(self):
        """Load sheets and columns from existing Excel file"""
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            
            print(f"\n✓ Loaded Excel file: {self.excel_path}")
            print(f"✓ Found {len(workbook.sheetnames)} sheet(s)")
            
            # Ask if user wants to edit existing structure
            edit_choice = input("\nWould you like to edit the sheet structure? (yes/no): ").strip().lower()
            
            if edit_choice in ['yes', 'y']:
                self._edit_existing_file(workbook)
            else:
                # Just load existing structure
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                    columns = [col for col in first_row if col is not None]
                    
                    if columns:
                        display_columns = columns[:min(4, len(columns))]
                        self.sheets[sheet_name] = {
                            'columns': columns,
                            'display_columns': display_columns
                        }
                        print(f"  - {sheet_name}: {len(columns)} columns")
            
            if not self.sheets:
                print("\n⚠ Warning: No valid sheets found with columns.")
                self.excel_path = None
                
        except Exception as e:
            print(f"\n✗ Error loading file: {e}")
            self.excel_path = None
    
    def _edit_existing_file(self, workbook):
        """Edit existing Excel file structure"""
        print("\n" + "="*60)
        print("EDIT EXISTING FILE")
        print("="*60)
        
        print("\nCurrent sheets:")
        for idx, sheet_name in enumerate(workbook.sheetnames, 1):
            print(f"  {idx}. {sheet_name}")
        
        print("\nOptions:")
        print("  1. Keep all existing sheets")
        print("  2. Select specific sheets to use")
        print("  3. Add new sheets")
        
        choice = input("\nYour choice (1-3): ").strip()
        
        if choice == '1':
            # Keep all sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                columns = [col for col in first_row if col is not None]
                
                if columns:
                    display_columns = columns[:min(4, len(columns))]
                    self.sheets[sheet_name] = {
                        'columns': columns,
                        'display_columns': display_columns
                    }
        
        elif choice == '2':
            # Select specific sheets
            print("\nEnter sheet numbers to use (comma-separated, e.g., 1,3,4):")
            selection = input("Sheets: ").strip()
            
            try:
                indices = [int(x.strip()) for x in selection.split(',')]
                sheet_names = workbook.sheetnames
                
                for idx in indices:
                    if 1 <= idx <= len(sheet_names):
                        sheet_name = sheet_names[idx-1]
                        sheet = workbook[sheet_name]
                        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                        columns = [col for col in first_row if col is not None]
                        
                        if columns:
                            display_columns = columns[:min(4, len(columns))]
                            self.sheets[sheet_name] = {
                                'columns': columns,
                                'display_columns': display_columns
                            }
            except:
                print("⚠ Invalid input. Loading all sheets.")
                self._edit_existing_file(workbook)
        
        elif choice == '3':
            # Add new sheets
            num_new = input("\nHow many new sheets to add? (1-5): ").strip()
            try:
                num_new = int(num_new)
                if 1 <= num_new <= 5:
                    for i in range(num_new):
                        sheet_name = input(f"\nNew sheet {i+1} name: ").strip()
                        if sheet_name and sheet_name not in workbook.sheetnames:
                            columns = self._setup_columns_for_sheet(sheet_name)
                            display_columns = self._setup_display_columns(columns)
                            
                            # Create sheet in workbook
                            new_sheet = workbook.create_sheet(title=sheet_name)
                            new_sheet.append(columns)
                            
                            self.sheets[sheet_name] = {
                                'columns': columns,
                                'display_columns': display_columns
                            }
                    
                    # Save updated workbook
                    workbook.save(self.excel_path)
                    print(f"\n✓ Added {num_new} new sheet(s) and saved to {self.excel_path}")
            except:
                print("⚠ Invalid input.")


class PlaceholderEntry(ttk.Entry):
    """Entry widget with proper placeholder support"""
    def __init__(self, master, placeholder="", **kwargs):
        super().__init__(master, **kwargs)
        self.placeholder = placeholder
        self.placeholder_color = 'gray60'
        self.default_fg_color = 'black'
        
        self.bind("<FocusIn>", self._on_focus_in)
        self.bind("<FocusOut>", self._on_focus_out)
        
        self._show_placeholder()
    
    def _show_placeholder(self):
        if not self.get():
            self.insert(0, self.placeholder)
            self.config(foreground=self.placeholder_color)
    
    def _on_focus_in(self, event):
        if self.get() == self.placeholder:
            self.delete(0, tk.END)
            self.config(foreground=self.default_fg_color)
    
    def _on_focus_out(self, event):
        if not self.get():
            self._show_placeholder()
    
    def get_value(self):
        """Get actual value (empty string if placeholder)"""
        value = self.get()
        return "" if value == self.placeholder else value


class UpdateWindow(tk.Toplevel):
    def __init__(self, master, parent_window, selected_item, config, sheet_name, **kwargs):
        super().__init__(master, **kwargs)
        self.parent_window = parent_window
        self.selected_item = selected_item
        self.config = config
        self.sheet_name = sheet_name
        
        self.title(f'Update Data - {sheet_name}')
        self.resizable(False, False)
        self.geometry("450x550")
        
        # Get current values
        current_values = self.parent_window.whole_stored_data.item(selected_item)['values']
        
        self.entries = {}
        self._create_widgets(current_values)
    
    def _create_widgets(self, current_values):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        
        # Create scrollable frame
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        columns = self.config.sheets[self.sheet_name]['columns']
        
        # Create entry fields
        for idx, col in enumerate(columns):
            ttk.Label(scrollable_frame, text=f"{col}:").grid(
                row=idx, column=0, sticky="w", pady=5, padx=5
            )
            entry = ttk.Entry(scrollable_frame, width=40)
            entry.grid(row=idx, column=1, sticky="ew", pady=5, padx=5)
            entry.insert(0, current_values[idx] if idx < len(current_values) else "")
            self.entries[col] = entry
        
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=1)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Save", command=self.save_update).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.destroy).pack(side=tk.LEFT, padx=5)
    
    def save_update(self):
        columns = self.config.sheets[self.sheet_name]['columns']
        new_values = tuple(self.entries[col].get() for col in columns)
        
        try:
            # Update Treeview
            self.parent_window.whole_stored_data.item(self.selected_item, values=new_values)
            
            # Update Excel
            workbook = openpyxl.load_workbook(self.config.excel_path)
            sheet = workbook[self.sheet_name]
            
            item_index = self.parent_window.whole_stored_data.index(self.selected_item)
            excel_row = item_index + 2
            
            for col_idx, value in enumerate(new_values, start=1):
                sheet.cell(row=excel_row, column=col_idx, value=value)
            
            workbook.save(self.config.excel_path)
            messagebox.showinfo("Success", "Data updated successfully!")
            self.destroy()
            
        except FileNotFoundError:
            messagebox.showerror("Error", f"Excel file not found: {self.config.excel_path}")
        except PermissionError:
            messagebox.showerror("Error", "File is open in another program. Please close it and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update: {str(e)}")


class DataDisplayWindow(tk.Toplevel):
    def __init__(self, master, config, sheet_name, **kwargs):
        super().__init__(master, **kwargs)
        self.config = config
        self.sheet_name = sheet_name
        
        self.title(f"Full Data Display - {sheet_name}")
        self.geometry("1400x950")
        self.resizable(True, True)
        
        self._create_widgets()
        self.load_data()
    
    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=1)
        
        # Treeview with scrollbars
        tree_frame = ttk.Frame(main_frame)
        tree_frame.grid(row=0, column=0, sticky="nsew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        v_scroll = ttk.Scrollbar(tree_frame, orient="vertical")
        h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        columns = self.config.sheets[self.sheet_name]['columns']
        
        self.whole_stored_data = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            xscrollcommand=h_scroll.set,
            yscrollcommand=v_scroll.set
        )
        
        v_scroll.config(command=self.whole_stored_data.yview)
        h_scroll.config(command=self.whole_stored_data.xview)
        
        self.whole_stored_data.grid(row=0, column=0, sticky="nsew")
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll.grid(row=1, column=0, sticky="ew")
        
        # Set column headings
        for col in columns:
            self.whole_stored_data.heading(col, text=col)
            self.whole_stored_data.column(col, width=120)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, sticky="ew", pady=10)
        
        ttk.Button(button_frame, text="Update Selected", command=self.open_update_window).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Refresh", command=self.refresh_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Close", command=self.destroy).pack(side=tk.LEFT, padx=5)
    
    def load_data(self):
        try:
            workbook = openpyxl.load_workbook(self.config.excel_path)
            sheet = workbook[self.sheet_name]
            
            # Clear existing data
            for item in self.whole_stored_data.get_children():
                self.whole_stored_data.delete(item)
            
            # Load new data
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    self.whole_stored_data.insert("", tk.END, values=row)
                    
        except FileNotFoundError:
            messagebox.showerror("Error", f"Excel file not found: {self.config.excel_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {str(e)}")
    
    def refresh_data(self):
        """Reload data from Excel file"""
        self.load_data()
        messagebox.showinfo("Success", "Data refreshed!")
    
    def open_update_window(self):
        selected = self.whole_stored_data.selection()
        
        if not selected:
            messagebox.showwarning("Warning", "Please select a row to update!")
            return
        
        UpdateWindow(self, parent_window=self, selected_item=selected[0], 
                    config=self.config, sheet_name=self.sheet_name)


class SheetFrame(ttk.Frame):
    """Individual frame for each sheet"""
    def __init__(self, parent, config, sheet_name, **kwargs):
        super().__init__(parent, **kwargs)
        self.config = config
        self.sheet_name = sheet_name
        self.data_display_window = None
        
        self._setup_styles()
        self._create_widgets()
        self.load_data()
    
    def _setup_styles(self):
        style = ttk.Style()
        style.configure("Disabled.TButton", background="light gray")
        style.configure("Enabled.TButton", background="green")
    
    def _create_widgets(self):
        # Configure grid
        self.columnconfigure(0, weight=2)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)
        
        # Left side: Data entry
        self._create_data_entry_frame()
        
        # Right side: Preview
        self._create_preview_frame()
    
    def _create_data_entry_frame(self):
        entry_frame = ttk.LabelFrame(self, text=f"Data Entry - {self.sheet_name}", padding="10")
        entry_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        
        # Create scrollable frame
        canvas = tk.Canvas(entry_frame)
        scrollbar = ttk.Scrollbar(entry_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        self.entries = {}
        columns = self.config.sheets[self.sheet_name]['columns']
        
        for idx, col in enumerate(columns):
            ttk.Label(scrollable_frame, text=f"{col}:").grid(
                row=idx, column=0, sticky="w", pady=5, padx=5
            )
            entry = PlaceholderEntry(
                scrollable_frame,
                placeholder=f"Enter {col.lower()}",
                width=40
            )
            entry.grid(row=idx, column=1, sticky="ew", pady=5, padx=5)
            entry.bind('<KeyRelease>', self.check_fields)
            self.entries[col] = entry
        
        scrollable_frame.columnconfigure(1, weight=1)
        
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        entry_frame.columnconfigure(0, weight=1)
        entry_frame.rowconfigure(0, weight=1)
        
        # Buttons
        button_frame = ttk.Frame(entry_frame)
        button_frame.grid(row=1, column=0, columnspan=2, pady=10)
        
        self.submit_button = ttk.Button(
            button_frame,
            text="Submit",
            style="Disabled.TButton",
            command=self.submit
        )
        self.submit_button.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Clear", command=self.clear).pack(side=tk.LEFT, padx=5)
    
    def _create_preview_frame(self):
        display_columns = self.config.sheets[self.sheet_name]['display_columns']
        
        preview_frame = ttk.LabelFrame(self, text=f"Preview - {self.sheet_name}", padding="10")
        preview_frame.grid(row=0, column=1, sticky="nsew")
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)
        
        # Treeview
        tree_frame = ttk.Frame(preview_frame)
        tree_frame.grid(row=0, column=0, sticky="nsew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        v_scroll = ttk.Scrollbar(tree_frame, orient="vertical")
        
        self.stored_data = ttk.Treeview(
            tree_frame,
            columns=display_columns,
            show="headings",
            yscrollcommand=v_scroll.set
        )
        
        v_scroll.config(command=self.stored_data.yview)
        
        for col in display_columns:
            self.stored_data.heading(col, text=col)
            self.stored_data.column(col, width=100)
        
        self.stored_data.grid(row=0, column=0, sticky="nsew")
        v_scroll.grid(row=0, column=1, sticky="ns")
        
        # Full view button
        ttk.Button(
            preview_frame,
            text="Full View",
            command=self.full_data_viewer
        ).grid(row=1, column=0, sticky="ew", pady=(10, 0))
    
    def check_fields(self, event=None):
        """Enable submit button only when all fields have valid data"""
        all_filled = all(entry.get_value().strip() for entry in self.entries.values())
        
        if all_filled:
            self.submit_button.configure(style="Enabled.TButton", state="normal")
        else:
            self.submit_button.configure(style="Disabled.TButton", state="normal")
    
    def clear(self):
        """Clear all entry fields"""
        for entry in self.entries.values():
            entry.delete(0, tk.END)
            entry._show_placeholder()
        self.check_fields()
    
    def load_data(self):
        """Load preview data from Excel"""
        try:
            workbook = openpyxl.load_workbook(self.config.excel_path)
            sheet = workbook[self.sheet_name]
            
            columns = self.config.sheets[self.sheet_name]['columns']
            display_columns = self.config.sheets[self.sheet_name]['display_columns']
            
            # Clear existing
            for item in self.stored_data.get_children():
                self.stored_data.delete(item)
            
            # Load data (skip header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    # Show only display columns
                    display_data = [row[columns.index(col)] if columns.index(col) < len(row) else "" 
                                   for col in display_columns]
                    self.stored_data.insert("", tk.END, values=display_data)
                    
        except FileNotFoundError:
            messagebox.showerror("Error", f"Excel file not found: {self.config.excel_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {str(e)}")
    
    def submit(self):
        """Submit data to Excel file"""
        columns = self.config.sheets[self.sheet_name]['columns']
        row_values = [self.entries[col].get_value() for col in columns]
        
        if not all(row_values):
            messagebox.showwarning("Warning", "Please fill in all fields!")
            return
        
        try:
            workbook = openpyxl.load_workbook(self.config.excel_path)
            sheet = workbook[self.sheet_name]
            sheet.append(row_values)
            workbook.save(self.config.excel_path)
            
            messagebox.showinfo("Success", f"Data submitted to '{self.sheet_name}'!")
            
            # Refresh displays
            self.load_data()
            if self.data_display_window and self.data_display_window.winfo_exists():
                self.data_display_window.load_data()
            
            self.clear()
            
        except PermissionError:
            messagebox.showerror("Error", "File is open in another program. Please close it and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to submit data: {str(e)}")
    
    def full_data_viewer(self):
        """Open full data display window"""
        if self.data_display_window is None or not self.data_display_window.winfo_exists():
            self.data_display_window = DataDisplayWindow(self.winfo_toplevel(), self.config, self.sheet_name)
        else:
            self.data_display_window.lift()
            self.data_display_window.focus()


class Window:
    def __init__(self, root, config):
        self.root = root
        self.config = config
        
        self.root.title("Multi-Sheet Data Entry System")
        self.root.geometry("1400x950")
        self.root.resizable(True, True)
        
        self._create_widgets()
    
    def _create_widgets(self):
        # Main container
        main_container = ttk.Frame(self.root, padding="10")
        main_container.grid(row=0, column=0, sticky="nsew")
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_container.columnconfigure(0, weight=1)
        main_container.rowconfigure(1, weight=1)
        
        # Info bar at top
        info_frame = ttk.Frame(main_container)
        info_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(
            info_frame, 
            text=f"Excel File: {self.config.excel_path}",
            font=('Arial', 10, 'bold')
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(
            info_frame,
            text=f"Total Sheets: {len(self.config.sheets)}",
            font=('Arial', 10)
        ).pack(side=tk.LEFT, padx=20)
        
        # Notebook for multiple sheets
        if len(self.config.sheets) > 1:
            self.notebook = ttk.Notebook(main_container)
            self.notebook.grid(row=1, column=0, sticky="nsew")
            
            # Create tab for each sheet
            for sheet_name in self.config.sheets:
                sheet_frame = SheetFrame(self.notebook, self.config, sheet_name)
                self.notebook.add(sheet_frame, text=sheet_name)
        else:
            # Single sheet - no tabs needed
            sheet_name = list(self.config.sheets.keys())[0]
            sheet_frame = SheetFrame(main_container, self.config, sheet_name)
            sheet_frame.grid(row=1, column=0, sticky="nsew")


def main():
    print("\n" + "="*60)
    print("MULTI-SHEET DATA ENTRY SYSTEM - STARTUP")
    print("="*60)
    
    config = Config()
    
    if not config.excel_path or not config.sheets:
        print("\n✗ Setup incomplete or cancelled.")
        print("Application will exit.")
        input("\nPress Enter to exit...")
        return
    
    print("\n" + "="*60)
    print("STARTING APPLICATION...")
    print("="*60)
    print(f"Excel File: {config.excel_path}")
    print(f"Total Sheets: {len(config.sheets)}")
    for sheet_name, sheet_config in config.sheets.items():
        print(f"\n  Sheet: {sheet_name}")
        print(f"    - Columns: {len(sheet_config['columns'])}")
        print(f"    - Preview: {', '.join(sheet_config['display_columns'])}")
    print("="*60 + "\n")
    
    root = tk.Tk()
    app = Window(root, config)
    root.mainloop()


if __name__ == "__main__":
    main()